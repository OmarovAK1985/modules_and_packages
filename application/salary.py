import openpyxl


def calculate_salary(file, number_worker):
    count = 0
    salary = None
    file_excel = openpyxl.load_workbook(file)
    list_excel = file_excel['list']
    while count == 0:
        salary = list_excel.cell(column=3, row=number_worker + 1).value
        if salary is not None:
            break
        number_worker = int(input('Введите правильный номер работника, чтобы получить его заработную плату: '))
    if salary > 0:
        name_of_worker = list_excel.cell(column=1, row=number_worker+1).value
        position = list_excel.cell(column=2, row=number_worker+1).value
        pension_contributions = round(salary * 0.1, 2)
        medical_contributions = round(salary * 0.02, 2)
        income_tax = round((salary - pension_contributions - medical_contributions)*0.1, 2)
        print('____________________________')
        print(f'ФИО сотрудника {name_of_worker}, Должность {position}')
        print(f'Оклад - {salary}')
        print(f'Удержания: {pension_contributions + medical_contributions + income_tax}')
        print(f'\t В том числе: \n\t\t Пенсионные взносы: {pension_contributions}'
                                f'\n\t\t Взносы на медстрахование: {medical_contributions}'
                                f'\n\t\t Индивидуальный подоходный налог: {income_tax}'
              f'\n Заработная плата к выплате {salary - pension_contributions - medical_contributions - income_tax}')
        print('________________________')





