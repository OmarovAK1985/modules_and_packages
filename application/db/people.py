import openpyxl


def list_of_jobs(file):
    file_excel = openpyxl.open(file)
    list_excel = file_excel['list']
    count = 0
    print('Список сотрудников организации: ')
    for i in range(2, list_excel.max_row + 1):
        count = count + 1
        name = list_excel.cell(column=1, row=i).value
        position = list_excel.cell(column=2, row=i).value
        print(f'\t{count}. ФИО: {name}, Должность: {position}')


